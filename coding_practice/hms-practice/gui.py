import openpyxl
import tkinter as tk
import code as mycode
import os

counter = 0
workbook = openpyxl.Workbook()

def on_compute_button_click():
    number = mycode.transform(int(entry.get()))
    result.config(text=("The palindrome is: " + str(number.palindrome()) + ". Number of cycles: " + str(number.get_cycles())) )

def on_db_button_click():
    global counter
    global workbook
    counter = counter + 1
    
    sheet = workbook.active
    entry_val = int(entry.get())

    number = mycode.transform(entry_val)

    sheet.cell(row = counter,column = 1).value = entry_val
    sheet.cell(row = counter,column = 2).value = number.palindrome()
    sheet.cell(row = counter,column = 3).value = number.get_cycles()

    workbook.save("mydata.xlsx")

def on_display_button_click():
    os.system("start EXCEL.EXE mydata.xlsx")

# Create the main window
window = tk.Tk()

canvas = tk.Canvas(window, width=600, height=300)
canvas.pack()

result = tk.Label(text="Here your result will be shown") 
result.pack()

entry = tk.Entry(window) 
canvas.create_window(300, 140, window=entry)

# Create a button
button = tk.Button(window, text="Compute palindrome", command=on_compute_button_click)
button.pack()

button_database = tk.Button(window, text="Safe to database",command=on_db_button_click)
button_database.pack()

button_display = tk.Button(window,text="Display database",command=on_display_button_click)
button_display.pack()
# Start the GUI event loop
window.mainloop()